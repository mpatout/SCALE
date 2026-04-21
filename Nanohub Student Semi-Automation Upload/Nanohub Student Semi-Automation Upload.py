"""
SCALE Student Migration Script
===============================
Identifies students that need to be added to the system by comparing a master
student list against a current system data export, then produces a formatted
Excel import file ready for upload.

Dependencies
------------
    pip install pandas openpyxl numpy

Script Workflow
---------------
    1. Load the master student list from the configured Excel workbook
       (reads the "New Student List" tab).
    2. Prompt the user at runtime for the path to the current system data
       export (reads the "students" tab).
    3. Match records by normalized First Name + Last Name (+ DOB when
       available) and by email; suppress any student already in the system.
    4. Filter to "Current" and "No Mentor" status students only.
    5. Map matched fields to the system import template column schema,
       applying value translations (vertical, gender, degree type, dates).
    6. Write a multi-tab Excel output (Users, WorkExperience, Degrees,
       Mentoring) to the Outputs/ folder next to this script.

Arguments (interactive prompt at runtime)
-----------------------------------------
    SCALE_Student_Data file path -- Full path to the system export
                                    (.xlsx / .xls). Drag-and-drop into
                                    the terminal window is supported.

Configuration (edit before running)
-------------------------------------
    MASTER_LIST_PATH       -- Path to the master student list workbook.
    CURRENT_SCALE_SEMESTER -- Semester label written to the
                              studentScaleSemester import field.
"""

import re
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ---------------------------------------------------------------------------
# Configuration – update these values before each run
# ---------------------------------------------------------------------------

# Semester label written to the studentScaleSemester column in the import file.
CURRENT_SCALE_SEMESTER = "Spring 2026"


class StudentMigrationProcessor:
    """Process student data to identify new students and prepare import template."""
    
    def __init__(self, master_list_path, system_data_folder=None, system_data_file=None):
        """
        Initialize the processor.
        
        Args:
            master_list_path: Path to the Updated Student List Excel file
            system_data_folder: Folder containing the downloaded SCALE_Student_Data file (optional)
            system_data_file: Direct path to a specific SCALE_Student_Data file (optional)
        """
        self.master_list_path = Path(master_list_path)
        self.system_data_folder = Path(system_data_folder) if system_data_folder else None
        self.system_data_file = Path(system_data_file) if system_data_file else None
        self.master_df = None
        self.system_df = None
        self.new_students_df = None
        self.recent_suppressed_df = pd.DataFrame()
        
    def load_master_list(self):
        """Load the master student list workbook (reads the 'New Student List' tab)."""
        print(f"Loading master list from: {self.master_list_path}")
        try:
            # Read from "New Student List" tab
            self.master_df = pd.read_excel(self.master_list_path, sheet_name='New Student List', engine='openpyxl')
            print(f"✓ Loaded {len(self.master_df)} students from master list (tab: 'New Student List')")
            return True
        except Exception as e:
            print(f"✗ Error loading master list: {e}")
            return False
    
    def find_latest_system_data(self):
        """Find the most recent SCALE_Student_Data file in the folder."""
        if not self.system_data_folder:
            return None
            
        print(f"\nSearching for system data in: {self.system_data_folder}")
        
        # Pattern: SCALE_Student_Data*.xlsx
        pattern = "SCALE_Student_Data*.xlsx"
        files = list(self.system_data_folder.glob(pattern))
        
        if not files:
            print(f"✗ No files matching '{pattern}' found")
            return None
        
        # Sort by modification time (most recent first)
        files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        latest_file = files[0]
        
        mod_time = datetime.fromtimestamp(latest_file.stat().st_mtime)
        print(f"✓ Found latest file: {latest_file.name}")
        print(f"  Modified: {mod_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        return latest_file
    
    def load_system_data(self):
        """Load the current system data from the most recent export."""
        # Use manually specified file if provided
        if self.system_data_file:
            system_file = self.system_data_file
            print(f"\nLoading system data from specified file: {system_file.name}")
        else:
            # Otherwise, auto-find the latest file
            system_file = self.find_latest_system_data()
        
        if system_file is None:
            return False
        
        if not system_file.exists():
            print(f"✗ File not found: {system_file}")
            return False
        
        try:
            # Read from the "students" tab by name
            self.system_df = pd.read_excel(system_file, sheet_name='students')
            print(f"✓ Loaded {len(self.system_df)} students from system data (tab: 'students')")
            return True
        except Exception as e:
            print(f"✗ Error loading system data: {e}")
            return False
    
    def normalize_name(self, name):
        """Normalize name for comparison (lowercase and normalized whitespace)."""
        if pd.isna(name):
            return ""
        name = str(name).lower().strip()
        # Remove extra whitespace
        name = re.sub(r'\s+', ' ', name)
        return name
    
    def normalize_dob(self, dob):
        """Normalize date of birth for comparison."""
        if pd.isna(dob):
            return None
        
        # If it's already a datetime
        if isinstance(dob, pd.Timestamp) or isinstance(dob, datetime):
            return dob.date()
        
        # Try parsing string dates
        try:
            parsed = pd.to_datetime(dob)
            return parsed.date() if not pd.isna(parsed) else None
        except:
            return None
    
    def create_match_key(self, df, first_name_col, last_name_col, dob_col=None):
        """
        Create a matching key based on First Name + Last Name + DOB.
        
        Args:
            df: DataFrame
            first_name_col: Name of first name column
            last_name_col: Name of last name column
            dob_col: Name of date of birth column (optional)
        """
        df = df.copy()
        
        # Normalize names
        df['_first_normalized'] = df[first_name_col].apply(self.normalize_name)
        df['_last_normalized'] = df[last_name_col].apply(self.normalize_name)
        
        # Create base key with names
        df['_match_key'] = df['_first_normalized'] + '|' + df['_last_normalized']
        
        # Add DOB if available
        if dob_col and dob_col in df.columns:
            df['_dob_normalized'] = df[dob_col].apply(self.normalize_dob)
            df['_match_key'] = df['_match_key'] + '|' + df['_dob_normalized'].astype(str)
        
        return df
    
    def identify_new_students(self):
        """
        Identify students in master list that are NOT in the system.
        Uses normalized exact matching on First Name + Last Name (+ DOB when available),
        plus email matching when email columns are present.
        """
        print("\n" + "="*60)
        print("IDENTIFYING NEW STUDENTS")
        print("="*60)
        
        # Map column names from master list to standardized names
        # This mapping will need to be adjusted based on actual column names
        master_name_mapping = self._detect_column_mapping(self.master_df, 'master')
        system_name_mapping = self._detect_column_mapping(self.system_df, 'system')
        
        # Validate required columns exist
        if 'first_name' not in master_name_mapping or 'last_name' not in master_name_mapping:
            print("\n✗ Error: Could not detect first_name or last_name columns in master list")
            print("  Available columns:", ', '.join(self.master_df.columns[:10]))
            self.new_students_df = pd.DataFrame()
            return self.new_students_df
        
        if 'first_name' not in system_name_mapping or 'last_name' not in system_name_mapping:
            print("\n✗ Error: Could not detect first_name or last_name columns in system data")
            print("  Available columns:", ', '.join(self.system_df.columns[:10]))
            self.new_students_df = pd.DataFrame()
            return self.new_students_df
        
        # Create match keys - only use DOB if both lists have it
        master_dob = master_name_mapping.get('dob')
        system_dob = system_name_mapping.get('dob')
        
        # Use DOB only if BOTH files have it
        use_dob = master_dob and system_dob
        
        master_with_key = self.create_match_key(
            self.master_df,
            master_name_mapping['first_name'],
            master_name_mapping['last_name'],
            master_dob if use_dob else None
        )
        
        system_with_key = self.create_match_key(
            self.system_df,
            system_name_mapping['first_name'],
            system_name_mapping['last_name'],
            system_dob if use_dob else None
        )
        
        # Find students in master list that are NOT in system
        system_keys = set(system_with_key['_match_key'].dropna())
        name_in_system_mask = master_with_key['_match_key'].isin(system_keys)
        new_student_mask = ~name_in_system_mask
        
        # ALSO check email matching - if emails match, they're the same person
        # This is standalone evidence to exclude a record, even if names differ.
        master_email_col = master_name_mapping.get('email')
        
        # Collect all email-like columns from system data to maximize matches
        system_email_cols = [
            col for col in system_with_key.columns
            if 'email' in str(col).lower()
        ]
        
        # Fallback: if master email column wasn't detected, try any email-like column
        if not master_email_col:
            master_email_cols = [
                col for col in master_with_key.columns
                if 'email' in str(col).lower()
            ]
            master_email_col = master_email_cols[0] if master_email_cols else None
        
        email_in_system_mask = pd.Series(False, index=master_with_key.index)
        if master_email_col and system_email_cols:
            # Normalize master emails
            master_emails = master_with_key[master_email_col].astype(str).str.lower().str.strip()
            
            # Build a set of all emails in system across all email columns
            system_email_series = []
            for col in system_email_cols:
                col_values = system_with_key[col].dropna().astype(str).str.lower().str.strip()
                system_email_series.append(col_values)
            
            if system_email_series:
                system_email_set = set(
                    pd.concat(system_email_series)
                    .replace('', pd.NA)
                    .dropna()
                )
            else:
                system_email_set = set()
            
            # If email is in system, mark as existing (not new)
            email_in_system_mask = master_emails.isin(system_email_set)
            
            # Combine masks: new student = NOT in name keys AND NOT in email keys
            new_student_mask = new_student_mask & ~email_in_system_mask

        # Track students suppressed because they already matched in system
        suppressed_match_mask = name_in_system_mask | email_in_system_mask
            
        
        # Only include students with status 'Current' or no-mentor variants
        status_col = master_name_mapping.get('status')
        if status_col and status_col in master_with_key.columns:
            status_values = master_with_key[status_col].apply(self.normalize_student_status).astype(str).str.lower().str.strip()
            # Only keep Current or No Mentor after normalization
            allowed_mask = (
                (status_values == 'current') |
                (status_values == 'no mentor')
            )
            excluded_count = (~allowed_mask & new_student_mask).sum()
            if excluded_count > 0:
                new_student_mask = new_student_mask & allowed_mask

        self.new_students_df = master_with_key[new_student_mask].copy().reset_index(drop=True)

        # Build and print 3-week suppression warning table (terminal only)
        app_date_col = None
        for col in master_with_key.columns:
            if str(col).strip().lower() == 'application date':
                app_date_col = col
                break

        if not app_date_col:
            app_date_patterns = ['application date', 'applicationdate']
            for col in master_with_key.columns:
                if any(pattern in str(col).lower() for pattern in app_date_patterns):
                    app_date_col = col
                    break

        if app_date_col and 'first_name' in master_name_mapping and 'last_name' in master_name_mapping:
            application_dates = pd.to_datetime(master_with_key[app_date_col], errors='coerce')
            cutoff_date = pd.Timestamp.now().normalize() - pd.Timedelta(days=21)
            recent_mask = application_dates >= cutoff_date
            recent_suppressed_mask = suppressed_match_mask & recent_mask

            reason_values = np.where(
                email_in_system_mask & name_in_system_mask,
                'Email match; Name+DOB match',
                np.where(email_in_system_mask, 'Email match', 'Name+DOB match')
            )

            display_dates = application_dates.apply(
                lambda date_value: f"{date_value.month}/{date_value.day}/{date_value.year}" if pd.notna(date_value) else ''
            )

            report_df = pd.DataFrame({
                'Name': (
                    master_with_key[master_name_mapping['first_name']].fillna('').astype(str).str.strip() + ' ' +
                    master_with_key[master_name_mapping['last_name']].fillna('').astype(str).str.strip()
                ).str.strip(),
                'Email': (
                    master_with_key[master_email_col].fillna('').astype(str).str.strip()
                    if master_email_col and master_email_col in master_with_key.columns else ''
                ),
                'Application Date': display_dates,
                'Suppression Reason': reason_values,
                '_sort_date': application_dates
            })

            self.recent_suppressed_df = (
                report_df[recent_suppressed_mask]
                .sort_values(by=['_sort_date', 'Name'], ascending=[False, True], na_position='last')
                .drop(columns=['_sort_date'])
                .reset_index(drop=True)
            )

            print("\n" + "=" * 60)
            print("SUPPRESSED RECENTLY (LAST 3 WEEKS)")
            print("=" * 60)
            if len(self.recent_suppressed_df) > 0:
                print(self.recent_suppressed_df.to_string(index=False))
            else:
                print("No recent students were suppressed due to existing system matches.")
        else:
            self.recent_suppressed_df = pd.DataFrame(columns=['Name', 'Email', 'Application Date', 'Suppression Reason'])
            print("\nWarning: Could not build suppression report (missing Application Date or name columns).")
        
        print(f"\n✓ NEW students to add: {len(self.new_students_df)}")
        
        return self.new_students_df
    
    def _detect_column_mapping(self, df, list_type):
        """
        Auto-detect column names from the dataframe.
        
        Args:
            df: DataFrame to analyze
            list_type: 'master' or 'system'
        """
        columns = df.columns.tolist()
        mapping = {}
        
        # Common patterns for first name
        first_name_patterns = ['first name', 'firstname', 'first', 'fname']
        for col in columns:
            col_lower = str(col).lower().strip()
            if any(pattern in col_lower for pattern in first_name_patterns):
                mapping['first_name'] = col
                break
        
        # Common patterns for last name
        last_name_patterns = ['last name', 'lastname', 'last', 'lname', 'surname']
        for col in columns:
            col_lower = str(col).lower().strip()
            if any(pattern in col_lower for pattern in last_name_patterns):
                mapping['last_name'] = col
                break
        
        # Common patterns for DOB
        dob_patterns = ['dob', 'date of birth', 'birthdate', 'birth date']
        for col in columns:
            col_lower = str(col).lower().strip()
            if any(pattern in col_lower for pattern in dob_patterns):
                mapping['dob'] = col
                break
        
        # Email pattern
        email_patterns = ['email', 'e-mail', 'email address']
        for col in columns:
            col_lower = str(col).lower().strip()
            if any(pattern in col_lower for pattern in email_patterns):
                mapping['email'] = col
                break

        # Status pattern
        status_patterns = ['status', 'application status']
        for col in columns:
            col_lower = str(col).lower().strip()
            if any(pattern in col_lower for pattern in status_patterns):
                mapping['status'] = col
                break
        
        return mapping
    
    def translate_vertical_tab(self, value):
        """
        Translate vertical tab values from master sheet format to import format.
        
        Mapping:
        - System-on-Chip (SoC) → SoC
        - CSME (graduate students only) → SoC
        - Radiation Hardening (RH) → RH
        - Embedded Security Systems/Trusted AI (ESS/TAI) → TAI
        - Heterogeneous Integration and Advanced Packaging (HI/AP) → HI/AP
        """
        if pd.isna(value) or value == '':
            return ''
        
        value_str = str(value).strip()
        
        # Case-insensitive mapping
        translation_map = {
            'system-on-chip (soc)': 'SoC',
            'csme (graduate students only)': 'SoC',
            'radiation hardening (rh)': 'RH',
            'embedded security systems/trusted ai (ess/tai)': 'TAI',
            'heterogeneous integration and advanced packaging (hi/ap)': 'HI/AP'
        }
        
        # Try exact match first (case-insensitive)
        value_lower = value_str.lower()
        for key, mapped_value in translation_map.items():
            if value_lower == key:
                return mapped_value
        
        # Try partial match
        value_lower = value_str.lower()
        if 'soc' in value_lower or 'system-on-chip' in value_lower or 'csme' in value_lower:
            return 'SoC'
        elif 'rh' in value_lower or 'radiation hardening' in value_lower:
            return 'RH'
        elif 'tai' in value_lower or 'ess/tai' in value_lower or 'trusted ai' in value_lower:
            return 'TAI'
        elif 'hi/ap' in value_lower or 'heterogeneous integration' in value_lower:
            return 'HI/AP'
        
        # Return original value if no match found
        return value_str
    
    def translate_gender(self, value):
        """
        Translate gender values from master sheet format to import format.
        
        Mapping:
        - Man → Male
        - Woman → Female
        """
        if pd.isna(value) or value == '':
            return ''
        
        value_str = str(value).strip()
        value_lower = value_str.lower()
        
        # Case-insensitive mapping
        if value_lower == 'man':
            return 'Male'
        elif value_lower == 'woman':
            return 'Female'
        elif value_lower == 'male':
            return 'Male'
        elif value_lower == 'female':
            return 'Female'
        else:
            # Return original value if no match found
            return value_str

    def normalize_student_status(self, value):
        """Normalize status output to allowed values: Current or No Mentor."""
        if pd.isna(value) or value == '':
            return ''

        value_str = str(value).strip()
        value_lower = re.sub(r'\s+', ' ', value_str.lower())

        if value_lower == 'current':
            return 'Current'
        if value_lower == 'no mentor' or re.fullmatch(r'current\s*-\s*no mentor', value_lower):
            return 'No Mentor'

        return value_str
    
    def format_date(self, value):
        """
        Format as a true Date object so Excel recognizes it.
        """
        if pd.isna(value) or value == '':
            return ''
        
        try:
            # Convert to datetime if it's a string
            if isinstance(value, str):
                parsed_date = pd.to_datetime(value)
            else:
                parsed_date = value
            
            # Return a true Python date object instead of a string
            return parsed_date.date() if hasattr(parsed_date, 'date') else parsed_date
        except:
            # Return empty string if parsing fails
            return ''

    def format_date_mmddyyyy(self, value):
        """
        Format date as MM/DD/YYYY with leading zeros.
        Handles datetime objects, timestamps, and string dates.
        """
        if pd.isna(value) or value == '':
            return ''

        try:
            parsed_date = pd.to_datetime(value)
            return parsed_date.strftime('%m/%d/%Y')
        except:
            return ''

    def format_month_year(self, value):
        """
        Format date as Month YYYY (e.g., May 2026).
        Handles datetime objects, timestamps, and string dates.
        """
        if pd.isna(value) or value == '':
            return ''

        try:
            parsed_date = pd.to_datetime(value)
            return parsed_date.strftime('%B %Y')
        except:
            return ''

    def truncate_text(self, value, max_length=499):
        """Trim text exports to the allowed free-response length."""
        if pd.isna(value) or value == '':
            return ''

        return str(value).strip()[:max_length]

    def parse_graduation_date(self, value):
        """
        Parse graduation date values.

        Supports term formats like:
        - Spring '26 -> 05/15/2026
        - Fall '26 -> 12/15/2026
        - Summer '26 -> 08/15/2026

        Falls back to normal date parsing when a standard date is provided.
        Returns a python date object or None.
        """
        if pd.isna(value) or value == '':
            return None

        # Handle already-date values
        if isinstance(value, (pd.Timestamp, datetime)):
            return value.date()

        value_str = str(value).strip()
        if not value_str:
            return None

        # Normalize punctuation for parsing
        normalized = value_str.replace('’', "'").replace(',', ' ')

        # Handle term/year text like:
        # Spring '26, Spring 26, Spring 2026, Sp '26, Fa 2026, Su26
        match = re.search(
            r"\b(spring|summer|fall|sp|su|fa)\b\s*[\'\s]*(\d{2,4})[\']?",
            normalized,
            re.IGNORECASE
        )

        # Also allow year before term, e.g. 26 Spring
        if not match:
            match = re.search(
                r"[\'\s]*(\d{2,4})[\']?\s*\b(spring|summer|fall|sp|su|fa)\b",
                normalized,
                re.IGNORECASE
            )

        if match:
            if match.re.pattern.startswith(r"\b"):
                term = match.group(1).lower()
                year_text = match.group(2)
            else:
                year_text = match.group(1)
                term = match.group(2).lower()

            year = int(year_text)
            if year < 100:
                year += 2000

            if term in ['spring', 'sp']:
                return datetime(year, 5, 15).date()
            if term in ['fall', 'fa']:
                return datetime(year, 12, 15).date()
            if term in ['summer', 'su']:
                return datetime(year, 8, 15).date()

        # Fallback: try normal date parsing
        try:
            parsed_date = pd.to_datetime(value_str)
            if pd.isna(parsed_date):
                return None
            return parsed_date.date()
        except:
            return None

    def is_masters_degree_value(self, value):
        """Return True when source value indicates a master's-level student."""
        if pd.isna(value) or value == '':
            return False

        value_str = str(value).strip().lower()
        masters_patterns = [
            'masters graduate student', 'master graduate student',
            'ms', 'm.s', 'master', 'masters'
        ]
        return any(pattern in value_str for pattern in masters_patterns)

    def get_program_years_prior(self, degree_value):
        """Return years to subtract for startDate calculation."""
        if self.is_masters_degree_value(degree_value):
            return 2

        # BS and PHD default to 4 years per migration rules
        return 4

    def calculate_start_date_from_graduation(self, graduation_date, degree_value):
        """
        Back-calculate start date from graduation date and degree type rules.

        BS/PHD: 4 years prior
        Masters: 2 years prior

        Month/day mapping based on graduation term date:
        - May 15 graduation  -> Aug 15 start (years prior)
        - Dec 15 graduation  -> Aug 15 start (years prior)
        - Aug 15 graduation  -> May 15 start (years prior)
        """
        if graduation_date is None:
            return None

        years_prior = self.get_program_years_prior(degree_value)
        start_year = graduation_date.year - years_prior

        if graduation_date.month == 8:
            start_month, start_day = 5, 15
        else:
            start_month, start_day = 8, 15

        return datetime(start_year, start_month, start_day).date()

    def map_degree_type(self, value):
        """Map source degree/level text to allowed values: BS or PHD."""
        if pd.isna(value) or value == '':
            return 'BS'

        value_str = str(value).strip().lower()

        phd_patterns = [
            'ph.d. graduate student', 'ph.d graduate student',
            'phd graduate student', 'phd', 'ph.d', 'doctoral', 'doctorate'
        ]
        bs_patterns = [
            'first year undergraduate student',
            'second year undergraduate student',
            'third year undergraduate student',
            'fourth year undergraduate student',
            'fifth year or above undergraduate student',
            'undergrad', 'undergraduate', 'bs', 'b.s', 'bachelor'
        ]
        if any(pattern in value_str for pattern in phd_patterns):
            return 'PHD'
        if self.is_masters_degree_value(value_str):
            return 'PHD'
        if any(pattern in value_str for pattern in bs_patterns):
            return 'BS'

        return 'BS'

    def map_to_import_template(self):
        """
        Map the new students data to the import template format.
        """
        print("\n" + "="*60)
        print("MAPPING TO IMPORT TEMPLATE")
        print("="*60)
        
        if self.new_students_df is None or len(self.new_students_df) == 0:
            return pd.DataFrame()
        
        # Detect column mappings from the new_students_df (which has same columns as master_df)
        master_cols = self._detect_column_mapping(self.new_students_df, 'master')
        
        # Initialize the import template dataframe
        template_df = pd.DataFrame()
        
        # Map basic fields
        first_name_col = master_cols.get('first_name')
        template_df['firstName'] = self.new_students_df[first_name_col] if first_name_col and first_name_col in self.new_students_df.columns else ''
        
        last_name_col = master_cols.get('last_name')
        template_df['lastName'] = self.new_students_df[last_name_col] if last_name_col and last_name_col in self.new_students_df.columns else ''
        
        # Try to find middle name
        middle_patterns = ['middle name', 'middlename', 'middle']
        middle_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in middle_patterns):
                middle_col = col
                break
        template_df['middleName'] = self.new_students_df[middle_col] if middle_col and middle_col in self.new_students_df.columns else ''
        
        # Preferred name/title
        preferred_patterns = ['preferred name', 'preferredname', 'preferred']
        preferred_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in preferred_patterns):
                preferred_col = col
                break
        template_df['preferredNameOrTitle'] = self.new_students_df[preferred_col] if preferred_col and preferred_col in self.new_students_df.columns else ''
        
        # Email
        email_col = master_cols.get('email')
        template_df['email'] = self.new_students_df[email_col] if email_col and email_col in self.new_students_df.columns else ''
        
        # Vertical - try to map from "NEW Technical Area" or similar
        vertical_patterns = ['vertical', 'technical area', 'new technical area']
        vertical_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in vertical_patterns):
                vertical_col = col
                break
        
        # Apply vertical tab translation
        if vertical_col and vertical_col in self.new_students_df.columns:
            template_df['vertical'] = self.new_students_df[vertical_col].apply(self.translate_vertical_tab)
        else:
            template_df['vertical'] = ''
        
        # LinkedIn
        linkedin_patterns = ['linkedin', 'linkedinpage']
        linkedin_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in linkedin_patterns):
                linkedin_col = col
                break
        template_df['linkedInPage'] = self.new_students_df[linkedin_col] if linkedin_col and linkedin_col in self.new_students_df.columns else ''
        
        # Web page
        template_df['webPage'] = ''

        # Required fixed value
        template_df['agreedToDataCollection'] = 1
        
        # Map citizenship - convert to 1/0 format
        citizen_patterns = ['citizen', 'u.s. citizen', 'iscitizen']
        citizen_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in citizen_patterns):
                citizen_col = col
                break
        
        if citizen_col:
            # Convert Yes/No to 1/0
            template_df['isCitizen'] = self.new_students_df[citizen_col].apply(
                lambda x: 1 if str(x).lower() in ['yes', 'y', '1', 'true'] else 
                         (0 if str(x).lower() in ['no', 'n', '0', 'false'] else '')
            )
        else:
            template_df['isCitizen'] = ''
        
        # Race
        race_patterns = ['race', 'ethnicity']
        race_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in race_patterns):
                race_col = col
                break
        template_df['race'] = self.new_students_df[race_col] if race_col and race_col in self.new_students_df.columns else ''
        
        # Gender
        gender_patterns = ['gender', 'sex']
        gender_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in gender_patterns):
                gender_col = col
                break
        
        # Apply gender translation
        if gender_col and gender_col in self.new_students_df.columns:
            template_df['gender'] = self.new_students_df[gender_col].apply(self.translate_gender)
        else:
            template_df['gender'] = ''
        
        # Grade/level (required blank)
        template_df['grade'] = ''
        template_df['level'] = ''
        
        # Application date (master sheet column B: Application Date)
        app_date_col = None
        for col in self.new_students_df.columns:
            if str(col).strip().lower() == 'application date':
                app_date_col = col
                break

        if not app_date_col:
            app_date_patterns = ['application date', 'applicationdate']
            for col in self.new_students_df.columns:
                if any(pattern in str(col).lower() for pattern in app_date_patterns):
                    app_date_col = col
                    break
        
        # Apply date formatting (M/D/YYYY)
        if app_date_col and app_date_col in self.new_students_df.columns:
            template_df['scaleApplicationDate'] = self.new_students_df[app_date_col].apply(self.format_date)
        else:
            template_df['scaleApplicationDate'] = ''

        # Scale PP Application Date (column AY - same source as scaleApplicationDate, from column B)
        if app_date_col and app_date_col in self.new_students_df.columns:
            # Now uses format_date to pass the raw Python date object
            template_df['scaleppApplicationDate'] = self.new_students_df[app_date_col].apply(self.format_date)
        else:
            template_df['scaleppApplicationDate'] = ''

        # Student SCALE semester – set via CURRENT_SCALE_SEMESTER at module top
        template_df['studentScaleSemester'] = CURRENT_SCALE_SEMESTER
        
        # Status
        status_patterns = ['status', 'current status', 'student current status']
        status_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in status_patterns):
                status_col = col
                break
        if status_col and status_col in self.new_students_df.columns:
            template_df['studentCurrentStatus'] = self.new_students_df[status_col].apply(self.normalize_student_status)
        else:
            template_df['studentCurrentStatus'] = ''

        # Required fixed value
        template_df['seekingJobOffer'] = 1
        
        # Organization (university)
        org_patterns = ['university', 'organization', 'name of university', 'school']
        org_col = None
        for col in self.new_students_df.columns:
            if any(pattern in str(col).lower() for pattern in org_patterns):
                org_col = col
                break
        template_df['organization'] = self.new_students_df[org_col] if org_col and org_col in self.new_students_df.columns else ''
        
        # Required fixed value
        template_df['organizationType'] = 'Academic'
        
        # GPA, credit hours, and about-me are not imported; leave blank.
        template_df['creditHoursFromScaleOrDefense'] = ''
        template_df['aboutMe'] = ''
        
        # Initialise remaining template columns to empty string.
        # Excludes fields already assigned above and those set to fixed
        # values immediately below (holdSecurityClearance, recieveContacts,
        # grantsDataPermsission, scaleVersion).
        other_fields = [
            'favoriteActivity', 'bestSkills',
            'participatedInScaleEventOrMeeting', 'helpedWithScaleStudentActivities',
            'participatedInScaleResearch', 'researchCreditsEarned', 'presentedWork',
            'hasApplied', 'enrolledInCourse', 'participatedInDesignProject',
            'recievedAcademicCredit', 'paidByScale', 'interest', 'reviewCofirmation',
            'scaleSocialMediaMembership', 'scaleNanoHubAccess',
            'haveMemberOfScaleGroupOnNanoHub', 'usedScaleSocialMediaIntershipOrJobFair',
            'ifCreditsEarnedDescribeActivities', 'attendedOrientationMeeting',
            'viewedOrientationFile', 'technicalTraining', 'localTraining',
            'scaleppVertical', 'studentScaleppSemester'
        ]
        
        for field in other_fields:
            if field not in template_df.columns:
                template_df[field] = ''

        # Required fixed values
        template_df['holdSecurityClearance'] = 0
        template_df['recieveContacts'] = 1
        template_df['grantsDataPermsission'] = 1
        template_df['scaleVersion'] = 'SCALE'
        
        # Ensure column order matches template
        template_column_order = [
            'firstName', 'middleName', 'lastName', 'preferredNameOrTitle', 'email',
            'vertical', 'webPage', 'linkedInPage', 'agreedToDataCollection',
            'favoriteActivity', 'bestSkills', 'organization', 'organizationType',
            'race', 'isCitizen', 'gender', 'studentScaleSemester', 'scaleApplicationDate',
            'studentCurrentStatus', 'seekingJobOffer', 'grade',
            'participatedInScaleEventOrMeeting', 'helpedWithScaleStudentActivities',
            'participatedInScaleResearch', 'researchCreditsEarned', 'presentedWork',
            'hasApplied', 'enrolledInCourse', 'participatedInDesignProject',
            'recievedAcademicCredit', 'paidByScale', 'interest', 'holdSecurityClearance',
            'recieveContacts', 'grantsDataPermsission', 'reviewCofirmation',
            'scaleSocialMediaMembership', 'scaleNanoHubAccess',
            'haveMemberOfScaleGroupOnNanoHub', 'usedScaleSocialMediaIntershipOrJobFair',
            'creditHoursFromScaleOrDefense', 'ifCreditsEarnedDescribeActivities', 'level',
            'attendedOrientationMeeting', 'viewedOrientationFile', 'technicalTraining',
            'localTraining', 'aboutMe', 'scaleVersion', 'scaleppVertical',
            'scaleppApplicationDate', 'studentScaleppSemester'
        ]
        
        # Reorder columns
        template_df = template_df.reindex(columns=template_column_order, fill_value='')

        # Cap open-ended response fields to the import limit
        free_response_fields = [
            'favoriteActivity', 'bestSkills', 'interest',
            'ifCreditsEarnedDescribeActivities', 'aboutMe'
        ]
        for field in free_response_fields:
            if field in template_df.columns:
                template_df[field] = template_df[field].apply(lambda value: self.truncate_text(value, 499))
        
        print(f"✓ Mapped {len(template_df)} students to import template format")
        print(f"  Template has {len(template_df.columns)} columns")
        
        return template_df
    
    def prepare_work_experience_tab(self):
        """Prepare WorkExperience tab data for new students."""
        return pd.DataFrame(columns=[
            'studentEmail', 'employerName', 'employerType', 'jobType',
            'defenseRelated', 'startDate', 'endDate', 'hoursPerWeek', 'academicCredit'
        ])
    
    def prepare_degrees_tab(self):
        """Prepare Degrees tab data for new students."""
        if self.new_students_df is None or len(self.new_students_df) == 0:
            return pd.DataFrame(columns=[
                'studentEmail', 'university', 'degreeType', 'major', 'startDate', 'graduationDate'
            ])
        
        master_cols = self._detect_column_mapping(self.new_students_df, 'master')

        # Try to find a source column to infer degree type
        degree_type_col = None
        degree_type_patterns = [
            'degree type', 'degree', 'educational status', 'program level', 'program'
        ]
        for col in self.new_students_df.columns:
            col_lower = str(col).lower()
            if any(pattern in col_lower for pattern in degree_type_patterns):
                degree_type_col = col
                break

        if not degree_type_col:
            fallback_degree_patterns = ['grade', 'level']
            for col in self.new_students_df.columns:
                col_lower = str(col).lower()
                if any(pattern in col_lower for pattern in fallback_degree_patterns):
                    degree_type_col = col
                    break
        
        degrees_data = []
        
        for idx, row in self.new_students_df.iterrows():
            email = row.get(master_cols.get('email', ''), '')
            
            if not email or pd.isna(email):
                continue
            
            # Try to find university
            university = ''
            org_patterns = ['university', 'organization', 'name of university', 'school']
            for col in self.new_students_df.columns:
                if any(pattern in str(col).lower() for pattern in org_patterns):
                    university = row.get(col, '')
                    break
            
            # Try to find major
            major = ''
            major_patterns = ['major', 'field', 'discipline']
            for col in self.new_students_df.columns:
                if any(pattern in str(col).lower() for pattern in major_patterns):
                    major = row.get(col, '')
                    break
            
            # Try to find graduation date
            grad_date = ''
            grad_patterns = ['graduation date', 'graduation', 'expected graduation']
            for col in self.new_students_df.columns:
                if any(pattern in str(col).lower() for pattern in grad_patterns):
                    grad_date = row.get(col, '')
                    break

            # Try to find start date
            start_date = ''
            start_patterns = ['start date', 'program start', 'enrollment date', 'enrolled date']
            for col in self.new_students_df.columns:
                if any(pattern in str(col).lower() for pattern in start_patterns):
                    start_date = row.get(col, '')
                    break

            degree_type_source = row.get(degree_type_col, '') if degree_type_col else ''
            degree_type = self.map_degree_type(degree_type_source)

            graduation_date_obj = self.parse_graduation_date(grad_date)
            start_date_obj = self.calculate_start_date_from_graduation(graduation_date_obj, degree_type_source)

            formatted_grad_date = self.format_date(graduation_date_obj) if graduation_date_obj else ''
            if start_date_obj:
                formatted_start_date = self.format_date(start_date_obj)
            else:
                formatted_start_date = self.format_date(start_date)
            
            degrees_data.append({
                'studentEmail': email,
                'university': university,
                'degreeType': degree_type,
                'major': major,
                'startDate': formatted_start_date,
                'graduationDate': formatted_grad_date
            })
        
        degrees_df = pd.DataFrame(degrees_data)
        
        # Return empty dataframe with proper columns if no data
        if len(degrees_df) == 0:
            degrees_df = pd.DataFrame(columns=[
                'studentEmail', 'university', 'degreeType', 'major', 'startDate', 'graduationDate'
            ])
        
        return degrees_df
    
    def prepare_mentoring_tab(self):
        """Prepare Mentoring tab data for new students."""
        # Create empty template - mentoring relationships need to be established later
        mentoring_df = pd.DataFrame(columns=['studentEmail', 'mentorEmail'])
        
        return mentoring_df

    def _apply_users_tab_styling(self, worksheet):
        header_fill = PatternFill(fill_type='solid', start_color='000000', end_color='000000')
        orange_fill = PatternFill(fill_type='solid', start_color='ED7D31', end_color='ED7D31')
        even_fill = PatternFill(fill_type='solid', start_color='A6A6A6', end_color='A6A6A6')
        odd_fill = PatternFill(fill_type='solid', start_color='D9D9D9', end_color='D9D9D9')
        header_font = Font(color='FFFFFF', bold=True)
        body_font = Font(color='000000')
        white_side = Side(style='thin', color='FFFFFF')
        white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

        orange_columns = {1, 3, 5, 6, 12, 13, 14, 15, 16, 17, 18, 19, 20, 49, 51, 52}

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                if col_idx in orange_columns:
                    cell.fill = orange_fill
                else:
                    if row_idx == 1:
                        cell.fill = header_fill
                    elif row_idx % 2 == 0:
                        cell.fill = even_fill
                    else:
                        cell.fill = odd_fill

                cell.font = header_font if row_idx == 1 else body_font
                cell.border = white_border
                
                # Format columns scaleApplicationDate (18) and scaleppApplicationDate (51) as native Excel dates
                if col_idx in [18, 51] and row_idx > 1:
                    if cell.value != '':
                        cell.number_format = 'm/d/yyyy'

    def _apply_standard_tab_styling(self, worksheet):
        header_fill = PatternFill(fill_type='solid', start_color='000000', end_color='000000')
        even_fill = PatternFill(fill_type='solid', start_color='A6A6A6', end_color='A6A6A6')
        odd_fill = PatternFill(fill_type='solid', start_color='D9D9D9', end_color='D9D9D9')
        header_font = Font(color='FFFFFF', bold=True)
        body_font = Font(color='000000')
        white_side = Side(style='thin', color='FFFFFF')
        white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for row_idx in range(1, max_row + 1):
            fill = header_fill if row_idx == 1 else (even_fill if row_idx % 2 == 0 else odd_fill)
            font = header_font if row_idx == 1 else body_font

            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.font = font
                cell.border = white_border

    def _apply_degrees_tab_styling(self, worksheet):
        """Apply standard styling plus date formatting for Degrees tab."""
        self._apply_standard_tab_styling(worksheet)
        max_row = worksheet.max_row

        for row_idx in range(2, max_row + 1):
            # Column E (5) is startDate, Column F (6) is graduationDate
            for col_idx in [5, 6]:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value and cell.value != '':
                    cell.number_format = 'm/d/yyyy'
    
    def save_output(self, output_df, output_path=None):
        """Save the output to an Excel file with all required tabs."""
        if output_df is None or len(output_df) == 0:
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime("%m.%d.%Y %I.%M %p").lstrip('0')
            # Save to an Outputs/ folder beside this script
            output_dir = Path(__file__).parent / "Outputs"
            
            # Create directory if it doesn't exist
            output_dir.mkdir(parents=True, exist_ok=True)
            
            base_name = f"SCALE New Student Import {timestamp}"
            output_path = output_dir / f"{base_name}.xlsx"
            suffix = 1
            while output_path.exists():
                output_path = output_dir / f"{base_name} ({suffix}).xlsx"
                suffix += 1
        
        # Prepare all tabs
        print("\nPreparing Excel tabs...")
        work_exp_df = self.prepare_work_experience_tab()
        degrees_df = self.prepare_degrees_tab()
        mentoring_df = self.prepare_mentoring_tab()
        
        # Save as Excel file with all tabs
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                output_df.to_excel(writer, sheet_name='Users', index=False)
                work_exp_df.to_excel(writer, sheet_name='WorkExperience', index=False)
                degrees_df.to_excel(writer, sheet_name='Degrees', index=False)
                mentoring_df.to_excel(writer, sheet_name='Mentoring', index=False)

                users_ws = writer.sheets.get('Users')
                work_exp_ws = writer.sheets.get('WorkExperience')
                degrees_ws = writer.sheets.get('Degrees')
                mentoring_ws = writer.sheets.get('Mentoring')

                if users_ws is not None:
                    self._apply_users_tab_styling(users_ws)
                    users_ws.freeze_panes = 'A2'
                if work_exp_ws is not None:
                    self._apply_standard_tab_styling(work_exp_ws)
                    work_exp_ws.freeze_panes = 'A2'
                if degrees_ws is not None:
                    self._apply_degrees_tab_styling(degrees_ws)
                    degrees_ws.freeze_panes = 'A2'
                if mentoring_ws is not None:
                    self._apply_standard_tab_styling(mentoring_ws)
                    mentoring_ws.freeze_panes = 'A2'
        except PermissionError:
            print("\n✗ Permission denied while writing the output file.")
            print("  The file may already be open in Excel or locked by OneDrive.")
            print(f"  Close the file and re-run, or use a different name/location.")
            return None
        
        print(f"\n✓ Output saved to: {output_path}")
        print(f"  Tab 'Users': {len(output_df)} new students")
        print(f"  Tab 'WorkExperience': {len(work_exp_df)} records (template)")
        print(f"  Tab 'Degrees': {len(degrees_df)} records")
        print(f"  Tab 'Mentoring': {len(mentoring_df)} records (empty template)")
        
        return output_path
    
    def run(self):
        """Execute the full migration process."""
        print("="*60)
        print("SCALE STUDENT MIGRATION SCRIPT")
        print("="*60)
        print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # Step 1: Load master list
        if not self.load_master_list():
            print("\n✗ Failed to load master list. Exiting.")
            return None
        
        # Step 2: Load system data
        if not self.load_system_data():
            print("\n✗ Failed to load system data. Exiting.")
            return None
        
        # Step 3: Identify new students
        self.identify_new_students()
        
        # Step 4: Map to import template
        output_df = self.map_to_import_template()
        
        # Step 5: Save output
        output_file = self.save_output(output_df)
        
        print("\n" + "="*60)
        print("PROCESS COMPLETE")
        print("="*60)
        
        return output_file


def main():
    """Main entry point for the script."""
    
    # -----------------------------------------------------------------------
    # Configuration – set MASTER_LIST_PATH to your local copy of the master
    # student list workbook before running.
    # -----------------------------------------------------------------------
    MASTER_LIST_PATH = r"C:\path\to\your\Updated_Student_List.xlsm"
    
    # Prompt user for SCALE_Student_Data file
    print("Please provide the path to the SCALE_Student_Data Excel file.")
    print("(You can drag and drop the file into this window)\n")
    
    system_data_path = input("SCALE_Student_Data file path: ").strip()
    
    # Clean up the path (remove PowerShell artifacts from drag-and-drop)
    # Remove leading '& ' if present (PowerShell adds this)
    if system_data_path.startswith('& '):
        system_data_path = system_data_path[2:].strip()
    
    # Remove quotes (single or double)
    system_data_path = system_data_path.strip('"').strip("'").strip()
    
    if not system_data_path:
        print("\n✗ No file path provided. Exiting.")
        return
    
    # Validate file exists
    system_data_file = Path(system_data_path)
    if not system_data_file.exists():
        print(f"\n✗ File not found: {system_data_path}")
        print("Please check the path and try again.")
        return
    
    if not system_data_file.suffix in ['.xlsx', '.xls']:
        print(f"\n✗ Invalid file type: {system_data_file.suffix}")
        print("Please provide an Excel file (.xlsx or .xls)")
        return
    
    print(f"\n✓ Using file: {system_data_file.name}")
    
    # Create processor with manual file path
    processor = StudentMigrationProcessor(
        MASTER_LIST_PATH, 
        system_data_file=system_data_path
    )
    
    # Run the migration
    output_file = processor.run()
    
    if output_file:
        print(f"\n✓ SUCCESS: Import file created at:\n  {output_file}")
    else:
        print("\n✓ Process completed. No new students to add.")


if __name__ == "__main__":
    main()
